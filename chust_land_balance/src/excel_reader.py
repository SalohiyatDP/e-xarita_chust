# -*- coding: utf-8 -*-
"""
Excel contour-statement reader.

Reads the land-balance / contour Excel workbooks (.xls / .xlsx) that accompany
each massif geodatabase.  These workbooks carry per-contour or per-cadastre
attributes (legal document name/date/number, cadastre code, address,
specialization) that enrich the GIS-derived parcels.

Two back-ends are supported, tried in order:

1. ``xlrd`` (for .xls) / ``openpyxl`` (for .xlsx) - preferred, pure Python.
2. ``arcpy.conversion.ExcelToTable`` - used when neither library is present in
   the ArcGIS Python install; converts the sheet to a scratch table that is then
   read with a SearchCursor.

The reader is deliberately schema-agnostic: it detects the header row, then
exposes rows as dicts keyed by the (normalised) header text, plus helpers to
build a lookup index on a chosen column.
"""

from __future__ import unicode_literals

import os

try:
    import xlrd
    _HAS_XLRD = True
except ImportError:                       # pragma: no cover
    xlrd = None
    _HAS_XLRD = False

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:                       # pragma: no cover
    openpyxl = None
    _HAS_OPENPYXL = False

try:
    import arcpy
    _HAS_ARCPY = True
except ImportError:                       # pragma: no cover
    arcpy = None
    _HAS_ARCPY = False


def _norm(text):
    if text is None:
        return ""
    s = text if isinstance(text, type("")) else ("%s" % text)
    return " ".join(s.split()).strip().lower()


class ExcelSheet(object):
    """A single worksheet exposed as a list of header-keyed dict rows."""

    def __init__(self, name, header, rows):
        self.name = name
        self.header = header              # list of original header strings
        self._norm_header = [_norm(h) for h in header]
        self.rows = rows                  # list of dict (normalised-header -> value)

    def __len__(self):
        return len(self.rows)

    def find_column(self, *candidates):
        """Return the normalised header key matching any candidate, or None."""
        for cand in candidates:
            nc = _norm(cand)
            for h in self._norm_header:
                if h == nc or (nc and nc in h):
                    return h
        return None

    def index_by(self, *key_candidates):
        """Build {key_value: row} indexed by the first matching column."""
        col = self.find_column(*key_candidates)
        if not col:
            return {}
        index = {}
        for row in self.rows:
            key = row.get(col)
            if key in (None, ""):
                continue
            index[_norm(key)] = row
        return index


def _detect_header(matrix, max_scan=15):
    """Find the row index that looks like the header (most non-empty cells)."""
    best_idx, best_score = 0, -1
    for i, row in enumerate(matrix[:max_scan]):
        score = sum(1 for c in row if (c not in (None, "")))
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx


def _matrix_to_sheet(name, matrix):
    if not matrix:
        return ExcelSheet(name, [], [])
    hdr_idx = _detect_header(matrix)
    header = [("%s" % (c if c not in (None, "") else "col%d" % j))
              for j, c in enumerate(matrix[hdr_idx])]
    norm_header = [_norm(h) for h in header]
    rows = []
    for raw in matrix[hdr_idx + 1:]:
        if all(c in (None, "") for c in raw):
            continue
        row = {}
        for j, key in enumerate(norm_header):
            row[key] = raw[j] if j < len(raw) else None
        rows.append(row)
    return ExcelSheet(name, header, rows)


# ---------------------------------------------------------------------------
# back-ends
# ---------------------------------------------------------------------------
def _read_xls(path, sheet=None):
    book = xlrd.open_workbook(path)
    sheet_names = book.sheet_names()
    target = sheet if sheet in sheet_names else sheet_names[0]
    sh = book.sheet_by_name(target)
    matrix = []
    for r in range(sh.nrows):
        matrix.append([sh.cell_value(r, c) for c in range(sh.ncols)])
    return _matrix_to_sheet(target, matrix)


def _read_xlsx(path, sheet=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    target = sheet if sheet in wb.sheetnames else wb.sheetnames[0]
    ws = wb[target]
    matrix = [list(row) for row in ws.iter_rows(values_only=True)]
    return _matrix_to_sheet(target, matrix)


def _read_with_arcpy(path, sheet=None):    # pragma: no cover - needs arcpy
    scratch = arcpy.env.scratchGDB
    out_name = "xls_%d" % abs(hash(path))
    out_table = os.path.join(scratch, out_name)
    if arcpy.Exists(out_table):
        arcpy.Delete_management(out_table)
    kwargs = {}
    if sheet:
        kwargs["Sheet"] = sheet
    arcpy.conversion.ExcelToTable(path, out_table, **kwargs)
    fields = [f.name for f in arcpy.ListFields(out_table)
              if f.type not in ("OID", "Geometry")]
    rows = []
    with arcpy.da.SearchCursor(out_table, fields) as cur:
        for row in cur:
            rows.append(dict((_norm(fields[i]), row[i]) for i in range(len(fields))))
    return ExcelSheet(sheet or "Sheet1", fields, rows)


def read_workbook(path, sheet=None, logger=None):
    """Read a worksheet from ``path`` and return an :class:`ExcelSheet`.

    Raises nothing for a missing file - returns an empty sheet and logs a
    warning so the pipeline can degrade gracefully.
    """
    if not path or not os.path.exists(path):
        if logger:
            logger.warning("Excel workbook not found: %s", path)
        return ExcelSheet(sheet or "", [], [])

    low = path.lower()
    try:
        if low.endswith(".xlsx") and _HAS_OPENPYXL:
            return _read_xlsx(path, sheet)
        if low.endswith(".xls") and _HAS_XLRD:
            return _read_xls(path, sheet)
        if _HAS_XLRD and low.endswith(".xls"):
            return _read_xls(path, sheet)
        if _HAS_ARCPY:
            return _read_with_arcpy(path, sheet)
    except Exception as exc:                  # pragma: no cover
        if logger:
            logger.error("Failed to read workbook %s: %s", path, exc)

    if logger:
        logger.warning("No Excel back-end available for %s "
                       "(install xlrd / openpyxl).", path)
    return ExcelSheet(sheet or "", [], [])


def list_sheets(path):
    """Return the sheet names of a workbook (best effort)."""
    low = (path or "").lower()
    try:
        if low.endswith(".xlsx") and _HAS_OPENPYXL:
            wb = openpyxl.load_workbook(path, read_only=True)
            return list(wb.sheetnames)
        if low.endswith(".xls") and _HAS_XLRD:
            return list(xlrd.open_workbook(path).sheet_names())
    except Exception:
        pass
    return []
