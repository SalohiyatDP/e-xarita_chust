# -*- coding: utf-8 -*-
"""
Land-balance report generator.

Renders the SVOD (summary) and detailed balance tables to:

* **HTML**  - always produced, UTF-8, zero dependencies, prints cleanly to PDF
              from any browser; the universal fallback.
* **Excel** - .xlsx via openpyxl when available (else .xls via xlwt).
* **PDF**   - via reportlab when available, using a Cyrillic-capable TrueType
              font (Arial on Windows, DejaVuSans otherwise).

Every back-end degrades gracefully: a missing optional library is logged and
skipped, never fatal.  Pure data in -> files out; no arcpy dependency.
"""

from __future__ import unicode_literals

import io
import os

from config import settings
from config import land_categories as lc

# Optional back-ends -------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    _HAS_OPENPYXL = True
except ImportError:                       # pragma: no cover
    _HAS_OPENPYXL = False

try:
    import xlwt
    _HAS_XLWT = True
except ImportError:                       # pragma: no cover
    _HAS_XLWT = False

try:
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table as RLTable,
                                    TableStyle, Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    _HAS_REPORTLAB = True
except ImportError:                       # pragma: no cover
    _HAS_REPORTLAB = False


_FONT_CANDIDATES = [
    ("Arial", "C:/Windows/Fonts/arial.ttf"),
    ("ArialUni", "C:/Windows/Fonts/ARIALUNI.TTF"),
    ("Tahoma", "C:/Windows/Fonts/tahoma.ttf"),
    ("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ("DejaVuSans", "/usr/share/fonts/dejavu/DejaVuSans.ttf"),
]


def _register_cyrillic_font():
    """Register and return a TTF font name that can render Cyrillic, or None."""
    if not _HAS_REPORTLAB:
        return None
    for name, path in _FONT_CANDIDATES:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                return name
            except Exception:
                continue
    return None


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------
def _html_escape(text):
    s = u"%s" % (text if text is not None else "")
    return (s.replace(u"&", u"&amp;").replace(u"<", u"&lt;")
             .replace(u">", u"&gt;"))


def _color_hex(key):
    rgb = lc.CATEGORY_COLORS.get(key)
    if not rgb:
        return None
    return u"#%02x%02x%02x" % rgb


def _table_html(table):
    parts = [u'<h2>%s</h2>' % _html_escape(table.subtitle or "")]
    parts.append(u'<table>')
    # header
    parts.append(u'<thead><tr>')
    for col in table.columns:
        bg = _color_hex(col["key"])
        style = u' style="background:%s"' % bg if bg else u""
        parts.append(u'<th%s>%s</th>' % (style, _html_escape(col["label"])))
    parts.append(u'</tr></thead><tbody>')
    for i, row in enumerate(table.rows):
        cls = u' class="total"' if i == 0 else u""
        parts.append(u'<tr%s>' % cls)
        for col in table.columns:
            val = row.get(col["key"], "")
            if isinstance(val, float):
                val = u"%.*f" % (settings.AREA_DECIMALS, val)
            align = u"right" if col.get("kind") == "num" else u"left"
            parts.append(u'<td style="text-align:%s">%s</td>'
                         % (align, _html_escape(val)))
        parts.append(u'</tr>')
    parts.append(u'</tbody></table>')
    if table.footnote:
        parts.append(u'<p class="footnote">%s</p>' % _html_escape(table.footnote))
    return u"".join(parts)


_HTML_CSS = u"""
body{font-family:Arial,'DejaVu Sans',sans-serif;font-size:11px;margin:18px;color:#222;}
h1{font-size:16px;text-align:center;margin:6px 0;}
h2{font-size:13px;margin:14px 0 4px;}
table{border-collapse:collapse;width:100%;margin-bottom:10px;}
th,td{border:1px solid #888;padding:2px 4px;font-size:10px;}
th{background:#eee;text-align:center;vertical-align:middle;}
tr.total td{font-weight:bold;background:#f6f6e8;}
.footnote{font-style:italic;color:#555;font-size:10px;}
.meta{color:#555;font-size:10px;text-align:center;margin-bottom:8px;}
"""


def write_html(tables, out_path, title, meta=None):
    html = [u'<!DOCTYPE html><html><head><meta charset="utf-8">',
            u'<title>%s</title><style>%s</style></head><body>' % (
                _html_escape(title), _HTML_CSS),
            u'<h1>%s</h1>' % _html_escape(title)]
    if meta:
        html.append(u'<div class="meta">%s</div>' % _html_escape(meta))
    for t in tables:
        html.append(_table_html(t))
    html.append(u'</body></html>')
    with io.open(out_path, "w", encoding="utf-8") as fh:
        fh.write(u"".join(html))
    return out_path


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def write_excel(tables, out_path, title):
    if _HAS_OPENPYXL:
        return _write_xlsx(tables, out_path, title)
    if _HAS_XLWT:
        return _write_xls(tables, _swap_ext(out_path, ".xls"), title)
    return None


def _swap_ext(path, ext):
    base, _ = os.path.splitext(path)
    return base + ext


def _write_xlsx(tables, out_path, title):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for t in tables:
        ws = wb.create_sheet((t.subtitle or "Sheet")[:30] or "Sheet")
        ws.append([title])
        ws.append([t.subtitle or ""])
        ws.append([])
        header = [c["label"] for c in t.columns]
        ws.append(header)
        hdr_row = ws.max_row
        for j, col in enumerate(t.columns, start=1):
            cell = ws.cell(row=hdr_row, column=j)
            cell.font = bold
            cell.alignment = center
            cell.border = border
            rgb = lc.CATEGORY_COLORS.get(col["key"])
            if rgb:
                cell.fill = PatternFill("solid",
                                        fgColor="%02X%02X%02X" % rgb)
        for i, row in enumerate(t.rows):
            values = []
            for col in t.columns:
                v = row.get(col["key"], "")
                if isinstance(v, float):
                    v = round(v, settings.AREA_DECIMALS)
                values.append(v)
            ws.append(values)
            r = ws.max_row
            for j in range(1, len(t.columns) + 1):
                ws.cell(row=r, column=j).border = border
            if i == 0:
                for j in range(1, len(t.columns) + 1):
                    ws.cell(row=r, column=j).font = bold
        if t.footnote:
            ws.append([])
            ws.append([t.footnote])
    wb.save(out_path)
    return out_path


def _write_xls(tables, out_path, title):    # pragma: no cover - legacy path
    wb = xlwt.Workbook(encoding="utf-8")
    bold = xlwt.easyxf("font: bold on; align: wrap on, vert centre, horiz center")
    for t in tables:
        name = (t.subtitle or "Sheet")[:31] or "Sheet"
        ws = wb.add_sheet(name)
        ws.write(0, 0, title)
        ws.write(1, 0, t.subtitle or "")
        base = 3
        for j, col in enumerate(t.columns):
            ws.write(base, j, col["label"], bold)
        for i, row in enumerate(t.rows):
            for j, col in enumerate(t.columns):
                v = row.get(col["key"], "")
                if isinstance(v, float):
                    v = round(v, settings.AREA_DECIMALS)
                ws.write(base + 1 + i, j, v)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# PDF (reportlab)
# ---------------------------------------------------------------------------
def write_pdf(tables, out_path, title, meta=None, logger=None):
    if not _HAS_REPORTLAB:
        if logger:
            logger.warning("reportlab not installed - PDF report skipped "
                           "(HTML report can be printed to PDF instead).")
        return None
    font = _register_cyrillic_font()
    if not font:
        font = "Helvetica"
        if logger:
            logger.warning("No Cyrillic TTF font found - PDF text may not "
                           "render Uzbek Cyrillic correctly.")
    doc = SimpleDocTemplate(out_path, pagesize=landscape(A3),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"],
                                 fontName=font, fontSize=14, alignment=1)
    sub_style = ParagraphStyle("s", parent=styles["Heading2"],
                               fontName=font, fontSize=11)
    meta_style = ParagraphStyle("m", parent=styles["Normal"],
                                fontName=font, fontSize=9, alignment=1)
    foot_style = ParagraphStyle("f", parent=styles["Normal"],
                                fontName=font, fontSize=8)

    story = [Paragraph(_html_escape(title), title_style), Spacer(1, 4)]
    if meta:
        story.append(Paragraph(_html_escape(meta), meta_style))
        story.append(Spacer(1, 4))

    for t in tables:
        story.append(Paragraph(_html_escape(t.subtitle or ""), sub_style))
        story.append(Spacer(1, 2))
        story.append(_reportlab_table(t, font))
        if t.footnote:
            story.append(Spacer(1, 2))
            story.append(Paragraph(_html_escape(t.footnote), foot_style))
        story.append(Spacer(1, 6))

    doc.build(story)
    if logger:
        logger.info("Exported PDF report: %s", out_path)
    return out_path


def _reportlab_table(table, font):
    header = [_wrap(c["label"]) for c in table.columns]
    data = [header]
    for row in table.rows:
        line = []
        for col in table.columns:
            v = row.get(col["key"], "")
            if isinstance(v, float):
                v = u"%.*f" % (settings.AREA_DECIMALS, v)
            line.append(u"%s" % (v if v is not None else ""))
        data.append(line)

    rl = RLTable(data, repeatRows=1)
    style = [
        ("FONT", (0, 0), (-1, -1), font, 6),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BACKGROUND", (0, 1), (-1, 1), colors.Color(0.96, 0.96, 0.9)),
        ("FONT", (0, 1), (-1, 1), font, 6),
    ]
    # colour the header cell per category
    for j, col in enumerate(table.columns):
        rgb = lc.CATEGORY_COLORS.get(col["key"])
        if rgb:
            style.append(("BACKGROUND", (j, 0), (j, 0),
                          colors.Color(rgb[0] / 255.0, rgb[1] / 255.0, rgb[2] / 255.0)))
    rl.setStyle(TableStyle(style))
    return rl


def _wrap(text):
    """Insert soft breaks into long header labels so columns stay narrow."""
    s = u"%s" % (text or "")
    return s


# ---------------------------------------------------------------------------
# top-level
# ---------------------------------------------------------------------------
def generate_report(massif_balance, out_dir, basename, logger=None):
    """Build SVOD + detail tables and write every available report format."""
    from src import table_builder as tb
    if not os.path.isdir(out_dir):
        os.makedirs(out_dir)

    svod = tb.build_svod_table(massif_balance)
    detail = tb.build_detail_table(massif_balance)
    tables = [svod, detail]
    title = svod.title
    meta = u"%s | Хўжаликлар сони: %d | Контурлар: %d" % (
        _today(), massif_balance.farm_count, massif_balance.parcel_count)

    outputs = {}
    # HTML (always)
    html_path = os.path.join(out_dir, basename + "_REPORT.html")
    outputs["html_report"] = write_html(tables, html_path, title, meta)
    if logger:
        logger.info("Exported HTML report: %s", html_path)

    # Excel
    xlsx_path = os.path.join(out_dir, basename + "_REPORT.xlsx")
    xls_out = write_excel(tables, xlsx_path, title)
    if xls_out:
        outputs["excel_report"] = xls_out
        if logger:
            logger.info("Exported Excel report: %s", xls_out)
    elif logger:
        logger.warning("No Excel back-end (openpyxl/xlwt) - Excel report skipped.")

    # PDF
    pdf_path = os.path.join(out_dir, basename + "_REPORT.pdf")
    pdf_out = write_pdf(tables, pdf_path, title, meta, logger)
    if pdf_out:
        outputs["pdf_report"] = pdf_out

    return outputs


def _today():
    import time
    return time.strftime("%d.%m.%Y")
